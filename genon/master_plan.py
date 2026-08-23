#!/usr/bin/env python3
"""Genon step-2 master plan: annual budgets -> per-chapter allocation + drop floors.

Source: data/cloud/content/allocation_norms/ncf_chapterwise_period_allocation.xlsx
(budget sheet = founder's realistic annual budgets; Chapters sheet = effort weights).
Standard durations (HANDOVER Decision 2): 40 min for classes <= VII, 45 for VIII, 50 for IX.
Floor: unit-dropping begins when teacher_minutes / canonical_minutes < 0.6
(compression doctrine); floor_minutes = 0.6 x canonical_minutes.
"""
import json
import math

import openpyxl

from pathlib import Path
REPO = Path(__file__).resolve().parent.parent          # aruvi-saas/
NORMS = REPO / "data" / "cloud" / "content" / "allocation_norms"
WB = str(NORMS / "ncf_chapterwise_period_allocation.xlsx")
DROP_THRESHOLD = 0.6

# budget-sheet subject label -> repo subject key; Chapters-sheet label -> same key
SUBJECT_KEY = {
    "English": "english",
    "Mathematics": "mathematics",
    "Science": "science",
    "Social Sciences": "social_sciences",
    "TWAU": "the_world_around_us",
    "The World Around Us": "the_world_around_us",
}
ROMAN = {"III": 3, "IV": 4, "V": 5, "VI": 6, "VII": 7, "VIII": 8, "IX": 9, "X": 10}


# How a subject·stage is SERVED decides how densely its canonicals must be authored.
#   "unit" — the standard engine: it fills a gap by borrowing a UNIT, so the library
#            only needs the equal-dispersion spread (architecture §0.2).
#   "plan" — science·middle alone: units belong to a cognitive progression arc, a stage
#            is taught whole or not at all, so no prefix of a canonical is a valid plan.
#            Serving is whole-canonical selection; the ONLY bridge between two counts is
#            the top's single synthesis unit, which spans a gap of exactly 1. So the
#            counts must step by 2 or the band has holes.
#            Spec: docs/science_middle_stage_serve.md §3 (founder, 2026-08-07).
# One table, read in one place — never an `if subject == …` in the body (CLAUDE.md §3).
SERVE_GRANULARITY = {("science", "middle"): "plan"}


def _stage_of(cls_roman):
    n = ROMAN[cls_roman]
    return "preparatory" if n <= 5 else "middle" if n <= 8 else "secondary"


def canonical_periods(a, c, subject=None, cls_roman=None):
    """The chapter's canonical set over [floor C, standard A].

    UNIT-granularity stages (ten of eleven) — equal dispersion, architecture §0.2
    (v2.0, 2026-08-03). No solver, no sigma: {A, mid, C} with mid = ceil((A+C)/2)
    when the band is wide enough (A-C >= 4), {A, C} when the midpoint would sit
    adjacent to an endpoint, {A} alone on a degenerate band.

    PLAN-granularity stages (science·middle) — step down by exactly 2 from A. The
    spacing is forced by the serve law, not tuned: a gap of 2 is the widest the
    one-synthesis-unit bridge can cross, so any wider spacing leaves period counts
    unservable inside the band.

    THE FLOOR CONCESSION (founder, 2026-08-07). Where the step-2 chain's last value
    above the floor is exactly C+1, appending C off-chain buys nothing: C+1 was
    already reachable as C plus the borrowed synthesis. Taking ONE more step to C-1
    instead costs the same single canonical and extends the servable band one period
    LOWER — converting the X = C-1 request from a truncation with declared drops (a
    class that never reaches the arc's terminus) into a complete plan. Applies to 13
    of science·middle's 37 chapters at ₹0.

    GUARDED, because "one period" is modest at A=18 and absurd at A=3: the extra step
    is taken only if C-1 is still at least HALF the standard. Unguarded the rule emits
    a ONE-period canonical for VI ch 1 (A=3) and a 3-period arc at 43% for VII ch 1
    (A=7) — an arc needs sittings to be an arc. The guard blocks exactly those two and
    touches nothing else. Consequence to keep in view: the floor is no longer a flat
    0.6 — on the conceded chapters it lands at 0.50-0.56, always by exactly one period,
    and the LOWEST canonical is the one C8 and the human gate should read hardest.
    """
    a, c = int(a), int(c)
    if subject is not None and cls_roman is not None and \
            SERVE_GRANULARITY.get((subject, _stage_of(cls_roman))) == "plan":
        chain = list(range(a, c, -2))           # A, A-2, A-4, … all strictly above C
        if not chain:
            return [a]
        if chain[-1] == c + 1 and 2 * (c - 1) >= a:
            return chain + [c - 1]              # the concession, on-chain
        return chain + [c]                      # …else the floor itself, off-chain
    if c >= a - 1:
        return [a]
    if a - c < 4:
        return [a, c]
    return [a, (a + c + 1) // 2, c]


def _load_pins():
    """Per-chapter canonical_periods overrides (data/cloud/content/allocation_norms/
    canonical_period_pins.json). See that file's _meta for the reasoning.

    WHY THIS EXISTS (2026-08-07). A chapter whose library is already AUTHORED must not
    have its counts moved by a later change to the counts rule: the row would then
    disagree with the files on disk and certification's "library complete" check would
    fail a perfectly good library. The alternative to pinning is a metered regeneration.
    Found the hard way — a hand-edit of master_plan.json pinning science VIII ch 6 was
    silently wiped by the next regeneration within the hour, because THIS SCRIPT REBUILDS
    EVERY ROW FROM THE WORKBOOK. A pin has to live outside the generated file to survive."""
    try:
        with open(NORMS / "canonical_period_pins.json") as f:
            return (json.load(f) or {}).get("pins") or {}
    except FileNotFoundError:
        return {}


PINS = _load_pins()
PINS_USED = []


def pinned_or(subject, cls_roman, chapter, computed):
    p = PINS.get(f"{subject}|{cls_roman}|{chapter}")
    if not p or not p.get("counts"):
        return computed
    counts = [int(k) for k in p["counts"]]
    if counts != computed:
        PINS_USED.append((f"{subject} {cls_roman} ch{chapter}", computed, counts,
                          p.get("reason", "")))
    return counts


def std_duration(cls_roman):
    n = ROMAN[cls_roman]
    if n <= 7:
        return 40
    if n == 8:
        return 45
    return 50


def largest_remainder(budget, weights):
    total_w = sum(weights)
    exact = [budget * w / total_w for w in weights]
    base = [int(x) for x in exact]
    rem = budget - sum(base)
    order = sorted(range(len(exact)), key=lambda i: exact[i] - base[i], reverse=True)
    for i in order[:rem]:
        base[i] += 1
    assert sum(base) == budget
    return base, exact


wb = openpyxl.load_workbook(WB, data_only=True)

budgets = {}  # (subject_key, class_roman) -> annual periods
for r in wb["budget"].iter_rows(min_row=2, values_only=True):
    if r[0] is None:
        continue
    key = (SUBJECT_KEY[r[0]], r[1])
    assert key not in budgets, f"duplicate budget row {key}"
    budgets[key] = r[2]

chapters = {}  # (subject_key, class_roman) -> [(ch, title, weight)]
for r in wb["Chapters"].iter_rows(min_row=2, values_only=True):
    if r[0] is None:
        continue
    chapters.setdefault((SUBJECT_KEY[r[0]], r[1]), []).append((r[2], r[3], r[4]))

plan = {}
skipped = []
for key in sorted(chapters, key=lambda k: (k[0], ROMAN[k[1]])):
    subject, cls = key
    budget = budgets.get(key)
    if budget is None:
        skipped.append((subject, cls, "no annual budget in budget sheet"))
        continue
    chs = sorted(chapters[key])
    alloc, exact = largest_remainder(budget, [c[2] for c in chs])
    dur = std_duration(cls)
    rows = []
    for (ch, title, w), periods, ex in zip(chs, alloc, exact):
        canonical_min = periods * dur
        floor_min = DROP_THRESHOLD * canonical_min
        floor_periods = round(floor_min / dur)  # nearest, not ceil (founder, 2026-07-31)
        rows.append({
            "chapter": ch,
            "title": title,
            "weight": w,
            "exact_share": round(ex, 2),
            "recommended_periods": periods,
            "canonical_minutes": canonical_min,
            "floor_minutes": round(floor_min, 1),
            "floor_periods_at_standard": floor_periods,
            "canonical_periods": pinned_or(subject, cls, ch,
                                           canonical_periods(periods, floor_periods,
                                                             subject=subject,
                                                             cls_roman=cls)),
            "placeholder": "Placeholder" in str(title),
        })
    plan[f"{subject}|{cls}"] = {
        "subject": subject,
        "class": cls,
        "standard_duration_minutes": dur,
        "annual_budget_periods": budget,
        "total_effort_weight": sum(c[2] for c in chs),
        "n_chapters": len(chs),
        "chapters": rows,
    }

# budgets with no chapter data (e.g. class X)
for key, budget in sorted(budgets.items(), key=lambda kv: (kv[0][0], ROMAN[kv[0][1]])):
    if budget is not None and key not in chapters:
        skipped.append((key[0], key[1], f"budget {budget} but no chapters in workbook"))

out = {
    "_meta": {
        "generated": "2026-07-24",
        "source_workbook": "data/cloud/content/allocation_norms/ncf_chapterwise_period_allocation.xlsx",
        "standard_durations": {"<=VII": 40, "VIII": 45, "IX": 50},
        "allocation_method": "largest remainder over chapter effort weights (same as Allocate.jsx / api allocator)",
        "floor_definition": f"unit-dropping begins when teacher_minutes/canonical_minutes < {DROP_THRESHOLD}; "
                            f"floor_minutes = {DROP_THRESHOLD} x recommended_periods x standard_duration; "
                            "floor_periods_at_standard rounded to the NEAREST whole period (founder, 2026-07-31)",
        "canonical_periods": "the chapter's canonical set by EQUAL DISPERSION over "
                             "[floor, standard] (architecture §0.2, v2.0 2026-08-03): "
                             "{A, ceil((A+C)/2), C} when A-C >= 4, {A, C} when 1 < A-C < 4, "
                             "{A} otherwise. No solver, no sigma, no mandated closing spans "
                             "— canonicals are authored free; the standard alone carries "
                             "the synthesis-anchor mandate. EXCEPTION, science·middle: it "
                             "serves at PLAN granularity (a cognitive-progression stage is "
                             "taught whole or not at all, so no prefix of a canonical is a "
                             "valid plan), and its only bridge between two counts is the "
                             "top's single synthesis unit, which spans a gap of exactly 1. "
                             "Its counts therefore STEP DOWN BY 2 from the standard — forced "
                             "by the serve law, not tuned. FLOOR CONCESSION (founder, "
                             "2026-08-07): where the chain's last value above the floor C is "
                             "exactly C+1, the chain takes one more step to C-1 instead of "
                             "appending C off-chain — C+1 was already reachable as C plus the "
                             "borrowed synthesis, so the concession costs nothing and extends "
                             "the servable band one period lower. GUARDED: taken only if C-1 "
                             "is at least half the standard, which blocks a 1-period canonical "
                             "at A=3 and a 43%-of-standard arc at A=7. So floor_periods_at_"
                             "standard remains the 0.6 floor, and the LOWEST canonical may sit "
                             "one period below it (0.50-0.56 of standard on 13 of 37 chapters). "
                             "Spec: docs/science_middle_stage_serve.md §3.",
        "serve_granularity": {"science|middle": "plan", "_all_others": "unit"},
        "skipped": [f"{s} {c}: {why}" for s, c, why in skipped],
    },
    "combos": plan,
}
# WRITING IS GUARDED — an IMPORT MUST NOT REBUILD THE ARTEFACT (2026-08-16, ARV-D-164).
# This file's whole body used to run at import, master_plan.json write included. Anything
# that did `import master_plan` to reach a pure helper — `canonical_periods` is the obvious
# one — silently regenerated the plan from the workbook, which DROPS every `canonical_plan`
# annotation `variant_plans.py` had written. That is not a cosmetic loss: `briefs_for`
# refuses a provisional row, so the next compact wave refuses all 40 chapters with
# "Row is provisional", pointing at the standards rather than at the import that erased
# their annotation. Cost to recover is one free re-annotate; cost to DIAGNOSE is the
# expensive part, so the guard lives here rather than in a runbook note.
# The rebuild is now an explicit act: `python3 genon/master_plan.py`.
if __name__ == "__main__":
    with open(NORMS / "master_plan.json", "w") as f:
        json.dump(out, f, indent=2, ensure_ascii=False)
    print("wrote", (NORMS / "master_plan.json").relative_to(REPO),
          "— re-run `python3 genon/variant_plans.py` to restore canonical_plan annotations")

# ---- no derived md: master_plan.json is the single artifact (2026-07-31) ----
# The human-readable master_plan.md was RETIRED after it served a stale floor
# to the founder (ceil vs round, corrected the same day): a derived view that
# can drift from its source will eventually lie. To eyeball the plan, read the
# JSON fresh (python3 -m json.tool) or GET /subjects/{s}/{g}/chapters, which
# carries recommended_periods, floors, and canonical_plan per chapter.

    for _who, _computed, _pinned, _why in PINS_USED:
        print(f"PINNED  {_who}: rule says {_computed}, pinned to {_pinned}")
        print(f"        {_why[:150]}")
    if PINS_USED:
        print(f"({len(PINS_USED)} pin(s) applied from canonical_period_pins.json — remove a "
              f"pin once its chapter is regenerated to the current rule)\n")
    print("combos planned:", len(plan), "| skipped:", len(skipped))
    for s, c, why in skipped:
        print("  skipped:", s, c, "-", why)
    tot = sum(p["annual_budget_periods"] for p in plan.values())
    print("total annual periods across portfolio:", tot)
    ss9 = plan["social_sciences|IX"]["chapters"][4]
    print("spot check SS IX ch5:", ss9)
